from tkinter import*
import datetime as dt
import pandas as pd
import openpyxl as op
from tkinter import messagebox



window=Tk()


dic={
    'Date':[],
    'Hour':[],
    'Use':[]
}

#load past information

wb=op.load_workbook(r'insta.xlsx').active
lst1=[]
for i in range(2,wb.max_row+1):
  for j in range(1,wb.max_column+1):
    lst1.append(wb.cell(i,j).value)
  dic['Date'].append(lst1[0])
  dic['Hour'].append(lst1[1])
  dic['Use'].append(lst1[2])
  lst1.clear()  
    

def register_yes():
  d=dt.datetime.now()
  lst=str(d).split()
  
  if lst[0] not in dic['Date']:
    dic['Date'].append(lst[0])
    dic['Hour'].append(lst[1][0:8])
    dic['Use'].append('Yes')
    df=pd.DataFrame(dic)
    df.to_excel(r'insta.xlsx',index=False)
    messagebox.showinfo('Information','Registered successfully...')
  else:
    messagebox.showerror('Error405','Information was registered today!!') 

  window.quit()
  x=pd.read_excel(r'insta.xlsx')
  print(x)
  

def register_no():
  d=dt.datetime.now()
  lst=str(d).split()

  if lst[0] not in dic['Date']:
    dic['Date'].append(lst[0])
    dic['Hour'].append(lst[1][0:8])
    dic['Use'].append('No')
    df=pd.DataFrame(dic)
    df.to_excel(r'insta.xlsx',index=False)
    messagebox.showinfo('Information','Registered successfully...')
  else:
    messagebox.showerror('Error405','Information was registered today!!') 

  window.quit()
  x=pd.read_excel(r'insta.xlsx')
  print(x)
  
window.title('Alarm insta')
window.geometry('300x100')
window.resizable(width=False,height=False)
Label(window,text='Did you use instageram today?').pack()
fr=Frame(window)
fr.pack(pady=30)
btn_yes=Button(fr,text='YES',command=register_yes)
btn_yes.pack(side='right')
btn_no=Button(fr,text='NO',command=register_no)
btn_no.pack(side='left')

window.mainloop()
